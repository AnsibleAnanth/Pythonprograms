import psutil
import datetime
import time
import schedule
import openpyxl

pid = int(input("Enter the Process ID:"))


def warning():

    cpuusage = psutil.cpu_precent(interaval=1)

    if cpuusage > 50:
        print("Cpu Usage is above 50", cpuusage)

    memusage = psutil.virtual_memory().cpu_precent

    if memusage > 50:
        print("Memory Usage is above 50", memusage)


def monitor():

    time = datetime.datetime.now().strftime("%Y%m%d -%H:%M:%S")
    p = psutil.process(pid)
    cpu = p.cpu_precent(interaval=1) / psutil.cpu_count()
    memory_mb = p.memory_full_info().rss / (1024*1024)
    memory = p.memory_percent()
    path = r".\Monitor_Result.xlsx"
    file = openpyxl.load_workbook(path)
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row+1, value=time)
    sheet.cell(column=2, row=sheet.max_row, value=pid)
    sheet.cell(column=3, row=sheet.max_row, value=cpu)
    sheet.cell(column=4, row=sheet.max_row, value=memory_mb)
    sheet.cell(column=5, row=sheet.max_row, value=memory)
    file.save(path)


schedule.every(1).seconds.do(warning)
schedule.every(5).seconds.do(monitor)

while True:
    schedule.run_pending()
    time.sleep(1)
